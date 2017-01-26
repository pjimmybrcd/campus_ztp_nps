import json
from lib import actions, Excel_Reader
import openpyxl
import pymysql
import re

class GetInventoryAction(actions.SessionAction):
    def __init__(self, config):
        super(GetInventoryAction, self).__init__(config)
        self._connection = pymysql.connect(
                  host=self._db_addr, 
                  user=self._db_user,      
                  passwd=self._db_pass,  
                  db=self._db_name)   

    def run(self, filepath, sheetname, ap_name_column_name, mac_plus_three_column_name, base_mac_column_name):
        cursor =  self._connection.cursor()
        self.update(filepath, sheetname, ap_name_column_name, mac_plus_three_column_name, base_mac_column_name, cursor)
        cursor.close()
        self._connection.commit()
        self._connection.close()
        return (True, 'Success')


    def update(self, filepath, sheetname, ap_name_column_name, mac_plus_three_column_name, base_mac_column_name, cursor):
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.get_sheet_by_name(sheetname)
	
        ap_name_index = -1
        mac_index = -1
        base_mac_index = -1
        for row in ws.iter_rows():
                for index, cell in enumerate(row):
                        if(str(cell.value)==ap_name_column_name):
                                ap_name_index = index
                        if(str(cell.value)==base_mac_column_name):
                                base_mac_index = index
                        if(str(cell.value)==mac_plus_three_column_name):
                                mac_index = index
                break
        print(ap_name_index, mac_index, base_mac_index)
        for row in ws.iter_rows(row_offset=1):
                device, mac, base_mac = "NULL", "NULL", "NULL"
                for index, cell in enumerate(row):
                        #Gets the AP name from the excel sheet.
                        if index == ap_name_index and cell.value!=None:
                                device = str(cell.value)
			
                        #Validates the MAC + 3 if it is in correct format 1CB9.C43C.3B53
                        if index == mac_index and cell.value!=None:
                                value = str(cell.value)
                                regex = re.compile('([0-9a-fA-F]{4}\.[0-9a-fA-F]{4}\.[0-9a-fA-F]{4})')
                                match = regex.match(value)
                                if(match):
                                     mac = match.group(1)
                                regex = re.compile('([0-9a-fA-F]{2}:[0-9a-fA-F]{2}:[0-9a-fA-F]{2}:[0-9a-fA-F]{2}:[0-9a-fA-F]{2}:[0-9a-fA-F]{2})')
                                match = regex.match(value)
                                if(match):
                                     mac = match.group(1).split(":")
                                     if(len(mac)==6):
                                          mac = mac[0]+mac[1]+"."+mac[2]+mac[3]+"."+mac[4]+mac[5]

                                regex = re.compile('([0-9a-fA-F]{12})')
                                match = regex.match(value)
                                if(match):
                                     mac = match.group(1)
                                     mac = mac[0:4]+"."+mac[4:8]+"."+mac[8:12]

                        #Validates the Base MAC if it is in correct format 1C:B9:C4:3C:3B:50
                        if index == base_mac_index and cell.value!=None:
                                value = str(cell.value)
                                regex = re.compile('([0-9a-fA-F]{4}\.[0-9a-fA-F]{4}\.[0-9a-fA-F]{4})')
                                match = regex.match(value)
                                if(match):
                                     base_mac = re.sub('[.]', '', match.group(1))
                                     base_mac = base_mac[0:2] + ":" + base_mac[2:4] + ":" + base_mac[4:6] + ":" + base_mac[6:8] + ":" + base_mac[8:10] + ":" + base_mac[10:12]
                                
                                regex = re.compile('([0-9a-fA-F]{2}:[0-9a-fA-F]{2}:[0-9a-fA-F]{2}:[0-9a-fA-F]{2}:[0-9a-fA-F]{2}:[0-9a-fA-F]{2})')
                                match = regex.match(value)
                                if(match):
                                     base_mac = match.group(1)

                                regex = re.compile('([0-9a-fA-F]{12})')
                                match = regex.match(value)
                                if(match):
                                     value = match.group(1)
                                     base_mac = value[0:2] + ":" + value[2:4] + ":" + value[4:6] + ":" + value[6:8] + ":" + value[8:10] + ":" + value[10:12]
				


                if(device!="NULL" or mac!="NULL" or base_mac!="NULL"):
                        print(device, mac, base_mac)
                        sql = "select count(*) from authorized where mac='%s'" % (mac)
                        cursor.execute(sql)
                        count = cursor.fetchone()[0]
                        if(count == 0):
                                sql = 'insert into authorized (mac, base_mac, device) values("%s", "%s", "%s")' % (mac, base_mac, device)
                        else:
                                sql = 'update authorized set device="%s", base_mac="%s" where mac="%s"' % (device, base_mac, mac)
                        cursor.execute(sql)
