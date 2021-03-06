import json
from lib import actions, Excel_Reader, ztp_utils
import openpyxl
import pymysql
import re
import logging
from logging.handlers import RotatingFileHandler

class GetInventoryAction(actions.SessionAction):
    def __init__(self, config):
        super(GetInventoryAction, self).__init__(config)

        self._logger = logging.getLogger('cleanup_logger')
	self._logger.setLevel(logging.DEBUG)
        self._fileHandler = RotatingFileHandler("/var/log/campus_ztp_cleanuplog", mode='w', maxBytes=10*1024*1024, backupCount=2, encoding=None, delay=0)
        self._fileHandler.setLevel(logging.DEBUG)
        self._logger.addHandler(self._fileHandler)

        self._connection = pymysql.connect(
                  host=self._db_addr, 
                  user=self._db_user,      
                  passwd=self._db_pass,  
                  db=self._db_name)
        self._cursor =  self._connection.cursor()
        self._ruckus_controller_ip = self.config['ruckus_controller_ip']
        self._ruckus_session = None

        #Regex for IP Address
        self._ip_addr_regex = re.compile('(\d+\.\d+\.\d+\.\d+)')
        #Regex for port ex: 1/1/27
        self._port_regex = re.compile('(\d+\/\d+\/\d+)')
        #Regex for ICX Output: "2cc5.d321.b3b3 1/1/23    Dynamic   233"
        self._icx_output_regex = re.compile('([0-9a-fA-F]{4}\.[0-9a-fA-F]{4}\.[0-9a-fA-F]{4})(\s*)(\s\d+\/1\/\d+)(\s*)(Dynamic)(\s*)(233)')

    def run(self, filepath, sheetname, ip_column_name, switch_name_column_name):
        self._logger.info("***** Cleanup Action Initiated *****")
        self._logger.info("Attempting to SSH to Ruckus Controller on IP: '%s'" % self._ruckus_controller_ip)
        self._ruckus_session = ztp_utils.ruckus_controller_start_session(self._ruckus_controller_ip,
                                                self._ruckus_controller_username,
						self._ruckus_controller_password,
						self._ruckus_controller_enable_username,
						self._ruckus_controller_enable_password, "ssh")
        results = self.clean(filepath, sheetname, ip_column_name, switch_name_column_name)
        self._cursor.close()
        self._connection.commit()
        self._connection.close()
        for obj in results:
                print obj
                self._logger.info(obj)

        self._ruckus_session.logout()
        self._logger.info("Done")
        return (True, "Finished")

    def clean(self, filepath, sheetname, ip_column_name, switch_name_column_name):
        self._logger.info("Starting to parse the spreadsheet.")
        results = []

        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.get_sheet_by_name(sheetname)

        ip_index = -1
        name_index = -1
        for row in ws.iter_rows():
                for index, cell in enumerate(row):
                        if(str(cell.value)==ip_column_name):
                                ip_index = index
                        if(str(cell.value)==switch_name_column_name):
                                name_index = index
                break

        print("IP Address Column Index: " + str(ip_index), "Switch Name Column Index: " + str(name_index))
        #self._logger.info("IP Address Column Index: " + str(ip_index), "Switch Name Column Index: " + str(name_index))
        

        for row in ws.iter_rows(row_offset=1):
               ip, name = "NULL", "NULL"
               for index, cell in enumerate(row):
                       if index == ip_index and cell.value!=None:
                               ip = cell.value
                       if index == name_index and cell.value!=None:
                               name = cell.value

               if(self._ip_addr_regex.match(ip)==None or name=="NULL"):
                       self._logger.info("Error Invalid IP Address in file or invalid name, IP: '%s' and Name:'%s'" % (ip, name))
                       continue

               results.extend(self.start_icx_session(ip, name))
        return results

    def start_icx_session(self, ip, name):
        initial_command = "skip-page-display"
        ending_command = "page-display"
        command = "show mac-address | inc 233"

        self._logger.info("Starting session to ICX switch with ip: '%s' and name: '%s'" % (ip, name))
        icx_session = ztp_utils.start_session(ip, self._username, self._password,
                                              self._enable_username, self._enable_password, "ssh")

        self._logger.info("Sending command: '%s'" % initial_command)
        (success, output) = ztp_utils.send_commands_to_session(icx_session, initial_command, False)
        if(success == False):
                self._logger.info("Error in send Command: '%s' for Device: '%s', Results: '%s', '%s'." % (initial_command, ip, success, output))
                return (False, ip, "Failure, Error in send command")

        self._logger.info("Sending command: '%s'" % command)
        (success, output) = ztp_utils.send_commands_to_session(icx_session, command, False)
        if(success == False):
                self._logger.info("Error in send Command: '%s' for Device: '%s', Results: '%s', '%s'." % (command, ip, success, output))
                return (False, ip, "Failure, Error in send command")
        result = self.parse_output(icx_session, ip, name, output)

        self._logger.info("Sending command: '%s'" % ending_command)
        (success, output) = ztp_utils.send_commands_to_session(icx_session, ending_command, False)
        if(success == False):
                self._logger.info("Error in send Command: '%s' for Device: '%s', Results: '%s', '%s'." % (ending_command, ip, success, output))
                return (False, ip, "Failure, Error in send command")

        icx_session.logout();
        return result


    def parse_output(self, icx_session, switch_ip, switch_name, output):
        self._logger.info("Parsing output for ICX switch name: '%s', switch IP: '%s'" % (switch_name, switch_ip))
        self._logger.info("Output: '%s'" % (output))

        results = []
        if output==None or len(output)==0 or len(output[0]["output"])==0 :
                results.append(True, switch_ip, "No output");
                return results

        for line in output[0]["output"][0].splitlines():
                match = self._icx_output_regex.match(line.strip())
                if match:
                        self._logger.info("Regex successful match for output line: '%s'" % (line))
                        results.append(self.verify_and_update(icx_session, switch_ip, switch_name, match.group(1).strip(), match.group(3).strip()))
                else:
                        self._logger.info("Regex failed match for output line: '%s'" % (line))
        return results
        
    def verify_and_update(self, icx_session, switch_ip, switch_name, mac, port):
        self._logger.info("Verifying and updating db for ICX Switch IP: '%s', AP MAC address: '%s', ICX port: '%s'" % (switch_ip, mac, port))
        sql = "select count(*) from authorized where mac='%s'" % (mac)
        self._cursor.execute(sql)
        count = self._cursor.fetchone()[0]
        if count==0:
               #Mac address is not in the database so don't do anything.
               self._logger.info("Warning AP MAC address not in database for AP MAC: '%s', on ICX port: '%s', on ICX IP:'%s'" % (mac, port, switch_ip))
               return (True, switch_ip, "Warning AP MAC address not in database for AP MAC: '%s', on ICX port: '%s', on ICX IP:'%s'" % (mac, port, switch_ip))
        else:
               #Mac Address was found
               self._logger.info("Mac address found in database for AP MAC: '%s'" % (mac))

        sql = "select ip, port, base_mac, device, switch_name from authorized where mac='%s'" % (mac)
        self._cursor.execute(sql)
        row = self._cursor.fetchone()
        db_ip = row[0]
        db_port = row[1]
        db_base_mac = row[2]
        db_ap_name = row[3]
        db_switch_name = row[4]
        
        if(db_ip==None or db_port==None or db_ip==None or db_base_mac==None or db_ap_name==None or db_switch_name=="NULL" or self._ip_addr_regex.match(db_ip)==None or self._port_regex.match(db_port)==None):
               self._logger.info("Warning Database was missing information for AP MAC:'%s'." % (mac))
        elif(db_ip!=switch_ip or db_port!=port or db_switch_name!=switch_name):
               self._logger.info("Warning Database had invalid information for AP MAC:'%s'." % (mac))
               self._logger.info("Warning Database Information: Switch IP: '%s', Switch Name: '%s', Switch Port: '%s' for device with MAC: '%s'" % (db_ip, db_switch_name, db_port, mac))
               self._logger.info("Warning Switch Information: Switch IP: '%s', Switch IP: '%s', Switch Port: '%s' for device with MAC: '%s'" % (switch_ip, switch_name, port, mac))
        else:
               self._logger.info("Database information for AP MAC: '%s' is up-to-date." % (mac))

        #Updates the port name on the ICX
        (success, description) = self.icx_port_name_update(icx_session, port, db_ap_name)
        if(success==False):
               return (False, switch_ip, description, "AP MAC: '%s', on ICX port: '%s', on ICX IP:'%s'" % (mac, port, switch_ip))

        #Updates the DB
        sql = "update authorized set ip='%s', switch_name='%s', port='%s' where mac='%s'" % (switch_ip, switch_name, port, mac)
        self._cursor.execute(sql)
        self._connection.commit()
        self._logger.info("Database updated for AP MAC:'%s'." % (mac))

        #Updates the name on the ruckus controller
        (success, description) = self.ruckus_controller_update(switch_ip, switch_name, db_base_mac, db_ap_name, port)
        if(success==False):
               return (False, switch_ip, description, "AP MAC: '%s', on ICX port: '%s', on ICX IP:'%s'" % (mac, port, switch_ip))
        self._logger.info("Verify and update complete.")
        return (True, switch_ip, "Update Successful")

    def icx_port_name_update(self, icx_session, port, ap_name):
        self._logger.info("Updating ICX Port Name to be: '%s' for port: '%s'" % (ap_name, port))
        icx_port_name_command = "interface ethernet %s;port-name %s;" % (port, ap_name)
        (success, output) = ztp_utils.send_commands_to_session(icx_session, icx_port_name_command, True)
        self._logger.info("ICX Port Naming Result: '%s', Output: '%s'" % (success, output))
        return (success, "ICX Port Naming Result")

    """
    # Not needed currently.
    def icx_port_down_update(self, icx_session, port, ap_name=""):
        self._logger.info("Performing port down operation.")
        icx_port_down_command = "authentication;dot1x enable ethernet '%s';mac-authentication enable ethernet '%s';interface ethernet '%s';no dual-mode 233;no authentication auth-default-vlan 233;dot1x port-control auto;no port-name '%s';vlan 233;no tagged ethernet '%s'" % (port, port, port, ap_name, port)
        (success, output) = ztp_utils.send_commands_to_session(icx_session, icx_port_name_command, True)
        self._logger.info("ICX Port Down Result: '%s', Output: '%s'" % (success, output))
    """

    def ruckus_controller_update(self, switch_ip, switch_name, base_mac, ap_name, port):
        ruckus_command = "ap %s;name \"%s %s\";description \"%s %s %s %s\";end" % (base_mac, ap_name, switch_name, ap_name, switch_name, switch_ip, port)
        (success, output) = ztp_utils.send_commands_to_session(self._ruckus_session, ruckus_command, True)
        self._logger.info("Ruckus Controller Naming Result: '%s', Output: '%s'" % (success, output))
        return (success, "Ruckus Controller Naming Result")
               
